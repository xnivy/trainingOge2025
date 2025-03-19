from tkinter import messagebox
import openpyxl
from openpyxl.styles import Font
import pandas as pd
import matplotlib.pyplot as plt
from openpyxl import Workbook
from openpyxl.drawing.image import Image as ExcelImage
import os
from tkinter import *
import random
from tkinter import ttk
import numpy as np
import networkx as nx
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
from matplotlib.figure import Figure
from PIL import Image, ImageTk
from openpyxl.styles import Font, PatternFill



# ================== ГЕНЕРАТОРЫ ДАННЫХ ==================

def generate_task1():
    if random.choice([True, False]):
        # ========== Первый вариант задания  ==========
        while True:
            pages = random.randint(10, 50)
            lines = random.randint(30, 50)
            chars = random.randint(40, 60)
            total_bytes = pages * lines * chars * 2  # 2 байта на символ
            if total_bytes % 1024 == 0:
                size = total_bytes // 1024
                break
        return {
            "type": "numerical",
            "condition": (
                f"1. Статья, набранная на компьютере, содержит {pages} страниц, "
                f"на каждой странице {lines} строк, в каждой строке {chars} символов. "
                "В одном из представлений Unicode каждый символ кодируется 2 байтами. "
                "Определите информационный объем статьи в Кбайтах."
            ),
            "answer": str(size)
        }
    else:
        # ========== Второй вариант задания вариант ==========
        # Базовые предметы (40 вариантов)
        base_items = [
            "пуф", "стул", "диван", "кресло", "кровать", "софа", "табурет",
            "банкетка", "канапе", "лежак", "ширма", "трюмо", "комод", "сундук",
            "парта", "лавка", "качалка", "пуфик", "оттоман", "реклайнер",
            "подставка", "кабинка", "качалка", "лежанка", "подлокотник",
            "пуфарик", "скамейка", "кабриолет", "капелла", "манекен", "органайзер",
            "подголовник", "подножка", "подсвечник", "подставка", "прикроватник",
            "раскладуха", "ретро-кресло", "супер-пуф", "трансформер"
        ]
        
        # Длинные слова (11 символов) - 25 вариантов
        long_items = [
            "полукресло", "раскладушка", "трансформер", "многосекция",
            "гипер-пуф", "мегакресло", "ультрадиван", "суперкровать",
            "мультисофа", "псевдокомод", "квази-ширма", "неокомод",
            "интеркресло", "мета-лежак", "архи-банкет", "прото-софа",
            "кибер-табурет", "нано-канапе", "вирт-трюмо", "гекса-пуф",
            "орто-лежак", "поли-ширма", "квант-софа", "теле-кресло",
            "ай-диванчик"
        ]
        
        # Формируем список
        target_word = random.choice(long_items)
        selected_base = random.sample(base_items, k=random.randint(5, 9))
        all_items = selected_base + [target_word]
        random.shuffle(all_items)
        
        # Удаляем дубликаты
        all_items = list(dict.fromkeys(all_items))
        
        # Форматируем текст
        original_text = "Предметы мебели: " + ", ".join(all_items) + "."
        
        return {
            "type": "koi8",
            "condition": (
                f"1. В кодировке КОИ-8 (8 бит/символ). Ученица написала текст:\n"
                f"«{original_text}»\n"
                "Удалила название одного предмета, а также лишние запятую и пробел. "
                "Размер уменьшился на 13 байт. Какое название удалено?"
            ),
            "answer": target_word
        }

def generate_task2():
    codes = {'А': '.-', 'У': '..-', 'Ж': '...-', 'Х': '....'}
    message = ''.join(random.choices(list(codes.keys()), k=random.randint(8,12)))
    encoded = ''.join([codes[c] for c in message])
    return {
        "type": "morse",
        "condition": (
            "2. От разведчика была получена следующая шифрованная радиограмма:\n\n"
            "Код Морзе:\n"
            "А: .-\nУ: ..-\nЖ: ...-\nХ: ....\n\n"
            f"Закодированное сообщение: {encoded}\n\n"
            "Определите количество букв в исходной радиограмме."
        ),
        "encoded": encoded,
        "table": codes,
        "answer": str(len(message))
    }

def generate_task3():
    a = random.randint(5,15)
    b = random.randint(16,25)
    return {
        "type": "logic",
        "condition": (
            f"3. Напишите наибольшее целое число X, для которого истинно высказывание:\n"
            f"НЕ(X <= {a}) И НЕ(Y > {b})"
        ),
        "answer": str(a)
    }

def generate_task4():
    cities = ['А', 'Б', 'В', 'Г', 'Д']
    size = len(cities)
    matrix = np.zeros((size, size), dtype=int)
    
    # Генерируем базовый путь
    path = []
    current = 0
    for i in range(1, size):
        path.append((current, i))
        current = i
    
    # Заполняем базовый путь случайными значениями
    path_weights = [random.randint(1, 5) for _ in range(len(path))]
    for (i, j), weight in zip(path, path_weights):
        matrix[i][j] = matrix[j][i] = weight
    
    # Находим максимальный участок в базовом пути
    target_max = max(path_weights)
    max_edge_index = path_weights.index(target_max)
    
    # Добавляем альтернативные пути с бóльшими суммарными весами
    for i in range(size):
        for j in range(i+1, size):
            if matrix[i][j] == 0 and random.random() < 0.2:
                # Делаем вес альтернативного пути больше суммы базового
                alt_weight = sum(path_weights) + random.randint(1, 3)
                matrix[i][j] = matrix[j][i] = alt_weight
    
    # Гарантируем, что кратчайший путь останется через target_max
    matrix[path[max_edge_index]] = target_max
    matrix[path[max_edge_index][1], path[max_edge_index][0]] = target_max
    
    # Очищаем главную диагональ
    np.fill_diagonal(matrix, 0)
    
    return {
        "type": "matrix",
        "condition": (
            "4. В таблице указана протяженность дорог между пунктами:\n\n"
            "Укажите длину самого длинного участка кратчайшего пути от пункта А до пункта Д. "
            "Перемещаться можно только по дорогам, указанным в таблице:"
        ),
        "matrix": matrix,
        "answer": str(target_max)
    }

def generate_task5():
    while True:
        b = random.randint(2, 8)  # Ограничиваем диапазон для b
        start = random.randint(20, 40)
        
        # Вычисляем конечное значение с гарантией целочисленности
        current = start
        current += 3  # 1-я команда: +3
        current += 3  # 2-я команда: +3
        current = current // b  # 3-я команда: деление (целочисленное)
        current += 3  # 4-я команда: +3
        current += 3  # 5-я команда: +3
        
        # Убеждаемся, что end положительный и разумный
        if 5 <= current <= 15 and (start + 6) % b == 0:
            return {
                "type": "omega",
                "condition": (
                    f"5. У исполнителя Омега две команды:\n"
                    "1. прибавь 3\n"
                    "2. раздели на b\n\n"
                    f"Программа 11211 переводит число {start} в число {current}. Определите значение b."
                ),
                "answer": str(b)
            }
def generate_task6():
    pairs = [(9,9), (9,10), (8,5), (11,6), (-11,10), (-5,9), (-10,10), (4,5), (8,6)]
    yes_count = sum(1 for s,t in pairs if s < 9 or t < 9)
    return {
        "type": "program",
        "condition": (
            "6. Сколько раз программа выведет 'YES'?\n\n"
            "Исходные данные:\n"
            "s и t принимают значения:\n" + 
            '\n'.join([f"({s}, {t})" for s,t in pairs]) + 
            "\n\nПрограмма:\n"
            "if s < 9 or t < 9:\n    print('YES')\nelse:\n    print('NO')"
        ),
        "answer": str(yes_count)
    }

def generate_task7():
    # Определяем компоненты URL с их типами
    url_parts = [
        ('http://', 'protocol'),
        ('weather.info', 'domain'),
        ('/foto/', 'path'),
        ('2019/', 'subdir'),
        ('winter.jpg', 'file')
    ]
    
    # Перемешиваем части URL, но сохраняем порядок номеров 1-5
    shuffled_parts = url_parts.copy()
    random.shuffle(shuffled_parts)
    
    # Создаем список для отображения "номер → часть"
    numbered_parts = [(i+1, part[0]) for i, part in enumerate(shuffled_parts)]
    
    # Создаем правильный порядок типов
    correct_order = ['protocol', 'domain', 'path', 'subdir', 'file']
    
    # Формируем ответ
    answer_mapping = {part[1]: num for num, part in zip(range(1,6), shuffled_parts)}
    answer = ' '.join(str(answer_mapping[t]) for t in correct_order)

    return {
        "type": "url",
        "condition": (
            "7. Соберите URL из следующих компонентов:\n\n"
            "Доступные части:\n" +
            '\n'.join([f"{num}) {text}" for num, text in numbered_parts]) +
            "\n\nВ ответ запишите последовательность чисел через пробел, "
            "соответствующую правильному порядку URL"
        ),
        "answer": answer
    }

def generate_task8():
    both = random.randint(500,1000)
    A = random.randint(2000,3000)
    B = random.randint(1000,2000)
    return {
        "type": "search",
        "condition": (
            "8. Поисковые запросы:\n\n"
            f"Фрегат & Эсминец = {both}\n"
            f"Фрегат | Эсминец = {A}\n"
            f"Эсминец = {B}\n\n"
            "Сколько страниц (в тыс.) будет найдено по запросу 'Фрегат'?"
        ),
        "answer": str(A + both - B)
    }

def generate_task9():
    G = nx.DiGraph()
    cities = ['А', 'Б', 'В', 'Г', 'Д', 'Е', 'Ж', 'З', 'И']
    
    while True:
        G.clear()
        for i in range(len(cities)):
            for j in range(i+1, len(cities)):
                if random.random() < 0.35 and cities[j] > cities[i]:
                    G.add_edge(cities[i], cities[j])
        
        try:
            paths = list(nx.all_simple_paths(G, 'А', 'И'))
            if len(paths) > 0:
                break
        except:
            continue

    fig = Figure(figsize=(8, 6))
    ax = fig.add_subplot(111)
    
    # Фиксированное позиционирование узлов слева-направо
    pos = {
        'А': (0, 0),
        'Б': (1, 1), 'В': (1, -1),
        'Г': (2, 2), 'Д': (2, 0), 'Е': (2, -2),
        'Ж': (3, 1), 'З': (3, -1),
        'И': (4, 0)
    }
    
    # Рисуем граф
    nx.draw(
        G, pos, ax=ax, with_labels=True,
        node_size=1200, node_color='#E8F8FF',
        edge_color='gray', width=1.5,
        arrowsize=25, font_size=12,
        connectionstyle='arc3,rad=0.1'  # Для изогнутых стрелок
    )
    
    return {
        "type": "graph",
        "condition": "9. На рисунке — схема дорог, связывающих города А, Б, В, Г, Д, Е, Ж, З, И.\n"
            "По каждой дороге можно двигаться только в направлении стрелки.\n"
            "Сколько существует различных путей из города А в город И?", 
        "graph": fig,
        "answer": str(len(paths))
    }

def generate_task10():
    # Генерируем случайные числа для разных систем счисления
    octal_num1 = ''.join(str(random.randint(0, 7)) for _ in range(random.randint(2, 3)))
    octal_num2 = ''.join(str(random.randint(0, 7)) for _ in range(random.randint(2, 3)))
    binary_num = '1' + ''.join(str(random.randint(0, 1)) for _ in range(random.randint(3, 5)))
    
    # Конвертируем в десятичную систему
    numbers = [
        int(octal_num1, 8),
        int(octal_num2, 8),
        int(binary_num, 2)
    ]
    
    return {
        "type": "numsys",
        "condition": (
            "10. Среди приведенных ниже трех чисел, записанных в различных системах счисления, "
            "найдите минимальное и запишите его в ответе в десятичной системе счисления.\n\n"
            f"{octal_num1}₈ (восьмеричная)\n"
            f"{octal_num2}₈ (восьмеричная)\n"
            f"{binary_num}₂ (двоичная)\n\n"
            "В ответе запишите только число, основание системы счисления указывать не нужно."
        ),
        "answer": str(min(numbers))
    }
def generate_task11():
    # Случайный выбор одного из 5 вариантов задания
    task_variant = random.choice([1, 2, 3, 4, 5])
    
    if task_variant == 1:
        return {
            "type": "search_text",
            "condition": (
                "11. В одном из произведений И. С. Тургенева, текст которого приведен в подкаталоге Тургенев каталога DEMO-12, "
                "присутствует персонаж Базаров. С помощью поисковых средств операционной системы и текстового редактора "
                "выясните имя Базарова.\n\n"
                "Выполните задание, распаковав архив на своем компьютере.\n"
                "DEMO-12.rar"
            ),
            "answer": "Евгений"
        }
    elif task_variant == 2:
        return {
            "type": "search_text",
            "condition": (
                "11. В одном из стихотворений В. В. Маяковского, текст которого приведен в подкаталоге Маяковский каталога Поэзия, "
                "автор рассказывает о том, что прошел «тысячу Аркольских мостов». С помощью поисковых средств операционной системы "
                "и текстового редактора выясните, в каком месяце это произошло.\n\n"
                "Выполните задание, распаковав архив на своем компьютере.\n"
                "Поэзия.rar"
            ),
            "answer": "октябрь"
        }
    elif task_variant == 3:
        return {
            "type": "search_text",
            "condition": (
                "11. В одном из произведений В. В. Набокова, текст которого приведен в каталоге Набоков, "
                "присутствует персонаж с фамилией Ганин. С помощью поисковых средств операционной системы "
                "и текстового редактора выясните имя этого персонажа.\n\n"
                "Выполните задание, распаковав архив на своем компьютере.\n"
                "Набоков.rar"
            ),
            "answer": "Алексей"
        }
    elif task_variant == 4:
        return {
            "type": "search_text",
            "condition": (
                "11. В одном из произведений М. Ю. Лермонтова, текст которого приведен в подкаталоге каталога Проза, "
                "упоминается о Койшаурской долине. С помощью поисковых средств операционной системы и текстового редактора "
                "определите город, из которого был родом русский извозчик, первым спустившийся на повозке по дороге между пропастью и утесом.\n\n"
                "Выполните задание, распаковав архив на своем компьютере.\n"
                "Проза.rar"
            ),
            "answer": "Тифлис"
        }
    elif task_variant == 5:
        return {
            "type": "search_text",
            "condition": (
                "11. В одном из произведений Н. В. Гоголя, текст которого приведен в подкаталоге каталога Проза, "
                "есть персонаж с должностью смотритель училищ. С помощью поисковых средств операционной системы "
                "и текстового редактора или браузера выясните фамилию этого героя.\n\n"
                "Выполните задание, распаковав архив на своем компьютере.\n"
                "Проза.rar"
            ),
            "answer": "Хлопов"
        }
def generate_task12():
    variant = random.choice([1, 2, 3])
    
    if variant == 1:
        return {
            "type": "files",
            "condition": (
                "12. Сколько файлов с расширением .doc содержится в подкаталогах каталога Task12?\n"
                "В ответе укажите только число.\n\n"
                "Выполните задание, распаковав архив на своем компьютере.\n"
                "Task12.rar"
            ),
            "answer": "3"
        }
    elif variant == 2:
        return {
            "type": "files",
            "condition": (
                "12. Сколько файлов с расширением .txt содержится в подкаталогах каталога DEMO-12?\n"
                "В ответе укажите только число.\n\n"
                "Выполните задание, распаковав архив на своем компьютере.\n"
                "DEMO-12.rar"
            ),
            "answer": "33"
        }
    elif variant == 3:
        return {
            "type": "files",
            "condition": (
                "12. Сколько файлов с расширением .docx содержится в подкаталогах каталога DEMO-12?\n"
                "В ответе укажите только число.\n\n"
                "Выполните задание, распаковав архив на своем компьютере.\n"
                "DEMO-12.rar"
            ),
            "answer": "0"
        }

def generate_task13():
    variant = random.choice([1, 2, 3])
    
    if variant == 1:
        return {
            "type": "presentation",
            "condition": (
                "13. Выберите ОДНО из предложенных ниже заданий: 13.1 или 13.2.\n\n"
                "13.1 Используя информацию и иллюстративный материал, содержащийся в каталоге «Мышь полевка», "
                "создайте презентацию из трех слайдов на тему «Мышь полевка». Требования:\n"
                "- Краткие иллюстрированные сведения о внешнем виде, ареале обитания, образе жизни и рационе\n"
                "- Все слайды в едином стиле с заголовками\n"
                "- Форматы: *.odp, *.ppt, *.pptx\n\n"
                "13.2 Создайте документ по образцу с точным оформлением:\n"
                "- Шрифт 14pt, отступ первой строки 1см\n"
                "- Выравнивание по ширине, таблица по центру\n"
                "- Специальное форматирование текста\n"
                "- Форматы: *.odt, *.doc, *.docx\n\n"
                "Мышь полевка.rar\n"
            ),
            "image": "Вариант131.png",
            "answer": "OK"
        }
    elif variant == 2:
        return {
            "type": "presentation",
            "condition": (
                "13. Выберите ОДНО из предложенных ниже заданий: 13.1 или 13.2.\n\n"
                "13.1 Используя информацию и иллюстративный материал, содержащийся в каталоге «Летучая мышь», "
                "создайте презентацию из трех слайдов на тему «Летучая мышь». Требования:\n"
                "- Краткие иллюстрированные сведения о внешнем виде, ареале обитания, образе жизни и рационе\n"
                "- Все слайды в едином стиле с заголовками\n"
                "- Форматы: *.odp, *.ppt, *.pptx\n\n"
                "13.2 Создайте документ по образцу с точным оформлением:\n"
                "- Шрифт 14pt, отступ первой строки 1см\n"
                "- Выравнивание по ширине, таблица по центру\n"
                "- Специальное форматирование текста\n"
                "- Форматы: *.odt, *.doc, *.docx\n\n"
                "Летучая мышь.rar\n"
            ),
            "image": "Вариант132.png",
            "answer": "OK"
        }
    elif variant == 3:
        return {
            "type": "presentation",
            "condition": (
                "13. Выберите ОДНО из предложенных ниже заданий: 13.1 или 13.2.\n\n"
                "13.1 Используя информацию и иллюстративный материал из каталога DEMO-13, "
                "создайте презентацию на тему «Домовая мышь». Требования:\n"
                "- Сведения о внешнем виде, среде обитания, питании\n"
                "- Все слайды в едином стиле с заголовками\n"
                "- Форматы: *.odp, *.ppt, *.pptx\n\n"
                "13.2 Создайте документ по образцу с точным оформлением:\n"
                "- Шрифт 14pt, отступ первой строки 1см\n"
                "- Специальное выравнивание в таблицах\n"
                "- Сложное форматирование текста\n"
                "- Форматы: *.odt, *.doc, *.docx\n\n"
                "DEMO-13.rar\n"
            ),
            "image": "Вариант133.png",
            "answer": "OK"
        }

def generate_task14():
    # Генерация данных
    districts = ['СВ', 'Ю', 'В', 'С']
    subjects = ['математика', 'информатика', 'немецкий язык', 
               'обществознание', 'русский язык']
    
    data = {
        'округ': np.random.choice(districts, 1000),
        'фамилия': ['Ученик ' + str(i) for i in range(1, 1001)],
        'предмет': np.random.choice(subjects, 1000),
        'балл': np.random.randint(200, 600, 1000)
    }
    
    df = pd.DataFrame(data)
    
    # Сохранение в Excel
    filename = 'task14.xlsx'
    writer = pd.ExcelWriter(filename, engine='openpyxl')
    df.to_excel(writer, index=False)
    
    # Настройка стилей
    workbook = writer.book
    worksheet = writer.sheets['Sheet1']
    
    # Добавление инструкций
    worksheet['H1'] = 'Ответы:'
    worksheet['H1'].font = openpyxl.styles.Font(color="FFFFFF")  # Белый цвет текста
    worksheet['H2'] = '=COUNTIFS(A:A,"СВ",C:C,"математика")'
    worksheet['H2'].font = openpyxl.styles.Font(color="FFFFFF")  # Белый цвет текста
    worksheet['H3'] = '=ROUND(AVERAGEIF(A:A,"Ю",D:D),2)'
    worksheet['H3'].font = openpyxl.styles.Font(color="FFFFFF")  # Белый цвет текста
    
    # Убираем заливку ячеек
    worksheet['H1'].fill = openpyxl.styles.PatternFill(fill_type=None)  # Нет заливки
    worksheet['H2'].fill = openpyxl.styles.PatternFill(fill_type=None)  # Нет заливки
    worksheet['H3'].fill = openpyxl.styles.PatternFill(fill_type=None)  # Нет заливки
    
    writer.close()
    
    # Расчет правильных ответов
    answer1 = df[(df['округ'] == 'СВ') & (df['предмет'] == 'математика')].shape[0]
    answer2 = round(df[df['округ'] == 'Ю']['балл'].mean(), 2)
    
    return {
        "type": "spreadsheet",
        "condition": (
            "14. В электронную таблицу занесли данные о тестировании учеников. Ниже приведены первые пять строк таблицы:\n\n"
            "В столбце А записан округ, в котором учится ученик; в столбце В  — фамилия; в столбце С  — любимый предмет; в столбце D  — тестовый балл. Всего в электронную таблицу были занесены данные по 1000 ученикам.\n\n"
            "Выполните задание.\n\n"
            "1.  Сколько учеников в Северо-Восточном округе (СВ) выбрали в качестве любимого предмета математику? Ответ на этот вопрос запишите в ячейку F2 таблицы.\n\n"
            "2.  Каков средний тестовый балл у учеников Южного округа (Ю)? Ответ на этот вопрос запишите в ячейку Н3 таблицы с точностью два знака после запятой.\n\n"
            "3.  Постройте круговую диаграмму, отображающую соотношение числа участников, сдающих информатику, немецкий язык и обществознание. Левый верхний угол диаграммы разместите вблизи ячейки G6.\n\n"
            f"Файл с данными: {filename}"
        ),
        "file": filename,
        "answer": f"{answer1}, {answer2}"
    }
def generate_task15():
    answer = "вправо\nзакрасить\nпока сверху свободно\nвверх\nзакрасить\nкц"
    return {
        "type": "robot",
        "condition": (
            "15. Напишите алгоритм для Робота, чтобы закрасить нужные клетки.\n"
            "Пример ответа:\nвправо\nзакрасить\n..."
        ),
        "answer": answer
    }

def generate_task16():
    # Генерация случайных данных
    n = random.randint(5, 15)
    temps = [random.randint(-10, 20) for _ in range(n)]
    positive_temps = [t for t in temps if t > 0]
    avg = round(sum(positive_temps)/len(positive_temps), 1) if positive_temps else 0
    
    return {
        "type": "temperature",
        "condition": (
            "16. Ученики 4 класса вели дневники наблюдения за погодой и ежедневно записывали дневную температуру.\n\n"
            "Требуется:\n"
            "1. Найти среднюю температуру дней с температурой > 0°C\n"
            "2. Определить количество таких дней\n\n"
            "Программа получает на вход количество дней N (1 ≤ N ≤ 31), затем N целых чисел - температуры.\n"
            "Гарантируется, что хотя бы один день имеет температуру выше 0°C.\n\n"
        ),
        "input_data": [n] + temps,
        "output_data": [avg, len(positive_temps)]
    }

class Task16Window(Toplevel):
    def __init__(self, parent, task_data):
        super().__init__(parent)
        self.task_data = task_data
        self.title("Задание 16 - Анализ температур")
        self.geometry("800x700")
        self.configure(bg='white')

        # Основной контейнер с прокруткой
        self.main_canvas = Canvas(self, bg='white', highlightthickness=0)
        self.scrollbar = ttk.Scrollbar(self, orient="vertical", command=self.main_canvas.yview)
        self.scroll_frame = Frame(self.main_canvas, bg='white')

        # Настройка прокрутки колесиком мыши
        self.main_canvas.bind_all("<MouseWheel>", self.on_mouse_scroll)  # Добавлено!
        
        self.main_canvas.configure(yscrollcommand=self.scrollbar.set)
        self.main_canvas.pack(side=LEFT, fill=BOTH, expand=True)
        self.scrollbar.pack(side=RIGHT, fill=Y)
        self.main_canvas.create_window((0, 0), window=self.scroll_frame, anchor="nw")

        # Обработчик обновления области прокрутки
        self.scroll_frame.bind(
            "<Configure>",
            lambda e: self.main_canvas.configure(
                scrollregion=self.main_canvas.bbox("all")
            )
        )
        

        # Условие задачи
        Label(self.scroll_frame, 
             text=self.task_data['condition'],
             font=('Arial', 12),
             wraplength=700,
             justify=LEFT,
             bg='white').pack(pady=15, padx=20, anchor='w')

        # Таблица с данными
        self.table_frame = Frame(self.scroll_frame, bg='white')
        self.table_frame.pack(pady=10, padx=40)
        
        # Заголовки таблицы
        headers = ["Входные данные"]
        for col, header in enumerate(headers):
            Label(self.table_frame, text=header, 
                 font=('Arial', 12, 'bold'), 
                 borderwidth=1, relief='solid',
                 width=20, bg='#f0f0f0').grid(row=0, column=col, sticky='nsew')

        # Входные данные
        input_data = self.task_data['input_data']
        self.labels = []
        for i, val in enumerate(input_data):
            bg = '#ffffff' if i%2 == 0 else '#f8f8f8'
            lbl = Label(self.table_frame, text=str(val), 
                     borderwidth=1, relief='solid',
                     width=20, bg=bg)
            lbl.grid(row=i+1, column=0, sticky='nsew')
            self.labels.append(lbl)

        # Кнопка копирования всей таблицы
        Button(self.table_frame, 
              text="📋 Копировать все данные", 
              command=self.copy_all_data,
              bg='#5DADE2',
              fg='white').grid(row=len(input_data)+1, column=0, pady=10, sticky='ew')

        # Блок ответов
        answer_frame = Frame(self.scroll_frame, bg='white')
        answer_frame.pack(pady=20, padx=40)

        Label(answer_frame, text="Средняя температура:", 
             font=('Arial', 12), bg='white').grid(row=0, column=0)
        self.avg_entry = Entry(answer_frame, font=('Arial', 12), width=10)
        self.avg_entry.grid(row=0, column=1, padx=10)

        Label(answer_frame, text="Количество дней:", 
             font=('Arial', 12), bg='white').grid(row=1, column=0, pady=10)
        self.count_entry = Entry(answer_frame, font=('Arial', 12), width=10)
        self.count_entry.grid(row=1, column=1, padx=10)

        Button(answer_frame, text="Проверить", 
              command=self.check_answer,
              bg='#58D68D', fg='white').grid(row=2, columnspan=2, pady=15)

        # Контекстное меню
        self.context_menu = Menu(self, tearoff=0)
        self.context_menu.add_command(label="Копировать", command=self.copy_selected)
        
        # Привязка событий
        for lbl in self.labels:
            lbl.bind("<Button-3>", self.show_context_menu)
            lbl.bind("<Button-1>", self.select_cell)

    def on_mouse_scroll(self, event):
        """Обработчик колесика мыши"""
        self.main_canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")

    def copy_all_data(self):
        """Копирование всех данных таблицы"""
        data = [str(val) for val in self.task_data['input_data']]
        self.clipboard_clear()
        self.clipboard_append('\n'.join(data))
        messagebox.showinfo("Успех", "Все данные скопированы в буфер обмена!")

    def copy_selected(self):
        """Копирование выделенных ячеек"""
        try:
            selected_data = [
                lbl.cget("text") for lbl in self.labels 
                if lbl.cget("bg") == '#FFFF00'
            ]
            if not selected_data:
                raise ValueError("Не выбраны ячейки для копирования")
                
            self.clipboard_clear()
            self.clipboard_append('\n'.join(selected_data))
            messagebox.showinfo("Успех", "Выделенные данные скопированы!")
            
        except Exception as e:
            messagebox.showerror("Ошибка", str(e))

    def select_cell(self, event):
        """Выделение ячейки при клике"""
        widget = event.widget
        if widget.cget("bg") == '#FFFF00':
            widget.config(bg='#ffffff' if self.labels.index(widget)%2 == 0 else '#f8f8f8')
        else:
            widget.config(bg='#FFFF00')

    def show_context_menu(self, event):
        """Показ контекстного меню"""
        self.context_menu.post(event.x_root, event.y_root)

    def check_answer(self):
        """Проверка ответа (оригинальная логика)"""
        try:
            user_avg = float(self.avg_entry.get())
            user_count = int(self.count_entry.get())
            correct_avg = self.task_data['output_data'][0]
            correct_count = self.task_data['output_data'][1]

            result = []
            if abs(user_avg - correct_avg) < 0.1:
                result.append("✓ Средняя температура верна")
            else:
                result.append(f"✗ Средняя температура: {correct_avg}")

            if user_count == correct_count:
                result.append("✓ Количество дней верно")
            else:
                result.append(f"✗ Количество дней: {correct_count}")

            messagebox.showinfo("Результат", "\n".join(result))
        
        except ValueError:
            messagebox.showerror("Ошибка", "Некорректный формат ввода")

    def new_example(self):
        """Генерация нового примера"""
        self.task_data = generate_task16()
        self.destroy()
        Task16Window(self.master, self.task_data)

    def check_answer(self):
        try:
            user_avg = float(self.avg_entry.get())
            user_count = int(self.count_entry.get())
            correct_avg = self.task_data['output_data'][0]
            correct_count = self.task_data['output_data'][1]

            result = []
            if abs(user_avg - correct_avg) < 0.1:
                result.append("✓ Средняя температура верна")
            else:
                result.append(f"✗ Средняя температура: {correct_avg}")

            if user_count == correct_count:
                result.append("✓ Количество дней верно")
            else:
                result.append(f"✗ Количество дней: {correct_count}")

            messagebox.showinfo("Результат", "\n".join(result))
        
        except ValueError:
            messagebox.showerror("Ошибка", "Некорректный формат ввода")

    def new_example(self):
        self.task_data = generate_task16()
        self.destroy()
        Task16Window(self.master, self.task_data)

# ================== Построение 15 задания ==================
class RobotTaskWindow(Toplevel):
    def __init__(self, parent):
        super().__init__(parent)
        self.title("Задание 15 - Программирование Робота")
        self.geometry("1200x800")
        self.configure(bg='white')
        
        # Главный контейнер с прокруткой
        self.main_canvas = Canvas(self, bg='white', highlightthickness=0)
        self.scrollbar = ttk.Scrollbar(self, orient="vertical", command=self.main_canvas.yview)
        self.scrollable_frame = Frame(self.main_canvas, bg='white')

        # Настройка прокрутки
        self.main_canvas.configure(yscrollcommand=self.scrollbar.set)
        self.main_canvas.pack(side=LEFT, fill=BOTH, expand=True)
        self.scrollbar.pack(side=RIGHT, fill=Y)
        self.main_canvas.create_window((0, 0), window=self.scrollable_frame, anchor="nw")

        # Обработчик изменения размера
        self.scrollable_frame.bind("<Configure>", lambda e: self.main_canvas.configure(
            scrollregion=self.main_canvas.bbox("all")
        ))

        # Условие задачи
        condition_text = """ИСПОЛНИТЕЛЬ РОБОТ

            Требуется:
            1. Закрасить все клетки ниже горизонтальной стены и левее вертикальной
            2. Проходы в стенах должны остаться незакрашенными
            3. Робот не должен разрушиться

            Доступные команды:
            - вверх/вниз/влево/вправо
            - закрасить
            - условия: сверху/снизу/слева/справа свободно
            - циклы: нц пока <условие> ... кц"""

        # Верхняя панель с условием
        condition_frame = Frame(self.scrollable_frame, bg='white')
        Label(condition_frame, 
            text=condition_text,
            font=('Arial', 12), 
            justify=LEFT,
            bg='white',
            wraplength=1100
        ).pack(pady=15, padx=20, anchor='w')
        condition_frame.pack(fill=X)

        # Нижняя панель с редактором и лабиринтом
        content_frame = Frame(self.scrollable_frame, bg='white')
        
        # Холст для лабиринта
        self.grid_size = 25
        self.cell_size = 30
        self.canvas = Canvas(content_frame,
            width=self.grid_size*self.cell_size,
            height=self.grid_size*self.cell_size,
            bg='white',
            highlightthickness=0
        )
        self.canvas.pack(side=LEFT, padx=20, pady=10)

        # Панель управления
        control_frame = Frame(content_frame, bg='white')
        
        # Редактор кода с прокруткой
        editor_frame = Frame(control_frame, bg='white')
        self.code_editor = Text(editor_frame,
            width=55,
            height=20,
            font=('Courier New', 12),
            wrap=WORD
        )
        
        # Добавляем возможность копирования/вставки
        self.code_editor.bind("<Control-c>", lambda e: self.copy_text())
        self.code_editor.bind("<Control-v>", lambda e: self.paste_text())
        
        # Контекстное меню для копирования/вставки
        self.context_menu = Menu(self.code_editor, tearoff=0)
        self.context_menu.add_command(label="Копировать", command=self.copy_text)
        self.context_menu.add_command(label="Вставить", command=self.paste_text)
        self.code_editor.bind("<Button-3>", self.show_context_menu)
        
        scroll_editor = ttk.Scrollbar(editor_frame, command=self.code_editor.yview)
        self.code_editor.configure(yscrollcommand=scroll_editor.set)
        
        self.code_editor.pack(side=LEFT, fill=BOTH, expand=True)
        scroll_editor.pack(side=RIGHT, fill=Y)
        editor_frame.pack(fill=BOTH, expand=True)

        # Кнопки управления
        btn_frame = Frame(control_frame, bg='white')
        Button(btn_frame,
            text="▶ Запустить программу",
            command=self.execute_code,
            bg='#58D68D',
            fg='white',
            font=('Arial', 12)
        ).pack(pady=10, fill=X)

        Button(btn_frame,
            text="⟳ Сбросить состояние",
            command=self.reset_maze,
            bg='#EC7063',
            fg='white',
            font=('Arial', 12)
        ).pack(pady=10, fill=X)

        Button(btn_frame,
            text="🗑️ Очистить алгоритм",
            command=lambda: self.code_editor.delete("1.0", END),
            bg='#F4D03F',
            fg='black',
            font=('Arial', 12)
        ).pack(pady=10, fill=X)

        btn_frame.pack(fill=X)
        control_frame.pack(side=RIGHT, padx=20, pady=10, fill=BOTH, expand=True)
        content_frame.pack(fill=BOTH, expand=True)

        # Генерация лабиринта
        self.generate_maze()
        self.robot_pos = (2, self.grid_size//2 + 1)
        self.robot_path = set()
        self.draw_maze()

        # Привязка прокрутки колесом мыши
        self.main_canvas.bind_all("<MouseWheel>", self.on_mouse_scroll)

    def copy_text(self):
        """Копирование текста"""
        self.clipboard_clear()
        text = self.code_editor.get("sel.first", "sel.last")
        self.clipboard_append(text)

    def paste_text(self):
        """Вставка текста"""
        text = self.clipboard_get()
        self.code_editor.insert("insert", text)

    def show_context_menu(self, event):
        """Показ контекстного меню"""
        self.context_menu.post(event.x_root, event.y_root)

    def execute_code(self):
        """Выполнение программы робота"""
        try:
            self.robot_path = set()
            code = self.code_editor.get("1.0", END)
            self.execute_commands(code.split('\n'))
            self.check_solution()  # Проверка выполнения условий
        except Exception as e:
            # В случае ошибки сбрасываем положение робота
            # self.reset_maze()
            messagebox.showerror("Ошибка выполнения", str(e))


    def generate_maze(self):
        """Генерация стен с проходами"""
        self.walls = {'horizontal': [], 'vertical': []}

        
        # Горизонтальная стена
        h_y = self.grid_size//2
        h_start = 2
        h_end = self.grid_size - 5
        self.walls['horizontal'] = [(x, h_y) for x in range(h_start, h_end)]
        
        # Вертикальная стена
        v_x = h_end - 1
        v_start = h_y
        v_end = self.grid_size - 3
        self.walls['vertical'] = [(v_x, y) for y in range(v_start, v_end)]
        
        # Проходы
        self.passages = {
            'horizontal': random.randint(h_start + 2, h_end - 3),
            'vertical': random.randint(v_start + 2, v_end - 3)
        }

    def draw_maze(self):
        """Отрисовка лабиринта и робота"""
        self.canvas.delete("all")
        
        # Рисуем стены
        for x, y in self.walls['horizontal']:
            if x != self.passages['horizontal']:
                self.canvas.create_rectangle(
                    x*self.cell_size, y*self.cell_size,
                    (x+1)*self.cell_size, (y+1)*self.cell_size,
                    fill='#34495e', outline=''
                )
        
        for x, y in self.walls['vertical']:
            if y != self.passages['vertical']:
                self.canvas.create_rectangle(
                    x*self.cell_size, y*self.cell_size,
                    (x+1)*self.cell_size, (y+1)*self.cell_size,
                    fill='#34495e', outline=''
                )
        
        # Рисуем робота
        rx, ry = self.robot_pos
        self.canvas.create_oval(
            rx*self.cell_size + 5, ry*self.cell_size + 5,
            (rx+1)*self.cell_size - 5, (ry+1)*self.cell_size - 5,
            fill='#e74c3c', tags='robot'
        )
        
        # Рисуем закрашенные клетки
        for x, y in self.robot_path:
            self.canvas.create_rectangle(
                x*self.cell_size, y*self.cell_size,
                (x+1)*self.cell_size, (y+1)*self.cell_size,
                fill='#3498db', outline=''
            )

    def execute_code(self):
        """Выполнение программы робота"""
        self.robot_path = set()
        code = self.code_editor.get("1.0", END)
        self.execute_commands(code.split('\n'))
        self.check_solution()  # Проверка выполнения условий

    def execute_commands(self, lines):
        """Интерпретатор команд"""
        try:
            ptr = 0
            while ptr < len(lines):
                line = lines[ptr].strip()
                
                if not line:
                    ptr += 1
                    continue
                
                # Обработка циклов
                if line.startswith('нц пока'):
                    condition = line[7:].strip()
                    loop_body = []
                    ptr += 1
                    while ptr < len(lines) and lines[ptr].strip() != 'кц':
                        loop_body.append(lines[ptr].strip())
                        ptr += 1
                    self.handle_loop(condition, loop_body)
                
                # Обработка условий
                elif line.startswith('если'):
                    condition = line[4:].strip()
                    if_body = []
                    ptr += 1
                    while ptr < len(lines) and lines[ptr].strip() != 'все':
                        if_body.append(lines[ptr].strip())
                        ptr += 1
                    self.handle_condition(condition, if_body)
                
                # Обычные команды
                else:
                    self.process_command(line)
                
                ptr += 1
                self.update()
                self.after(150)

        except Exception as e:
            messagebox.showerror("Ошибка выполнения", str(e))

    def process_command(self, cmd):
        """Обработка одиночной команды"""
        x, y = self.robot_pos
        moves = {
            'вверх': (0, -1),
            'вниз': (0, 1),
            'влево': (-1, 0),
            'вправо': (1, 0),
            'закрасить': None
        }
        
        if cmd == 'закрасить':
            self.robot_path.add((x, y))
            return
        
        if cmd in moves:
            dx, dy = moves[cmd]
            new_x = x + dx
            new_y = y + dy
            if self.is_valid_move(new_x, new_y):
                self.robot_pos = (new_x, new_y)
            else:
                raise ValueError(f"Невозможно выполнить: {cmd}")

    def is_valid_move(self, x, y):
        """Проверка допустимости перемещения"""
        # Проверка горизонтальной стены
        if (x, y) in self.walls['horizontal'] and x != self.passages['horizontal']:
            return False
        # Проверка вертикальной стены
        if (x, y) in self.walls['vertical'] and y != self.passages['vertical']:
            return False
        return 0 <= x < self.grid_size and 0 <= y < self.grid_size

    def check_condition(self, condition):
        """Проверка условий с поддержкой отрицания"""
        x, y = self.robot_pos
        negate = False
        cond = condition.strip()

        # Если условие начинается с "не ", убираем его и помечаем инверсию
        if cond.startswith("не "):
            negate = True
            cond = cond[3:].strip()

        # Словарь направлений
        directions = {
            'сверху свободно': (0, -1),
            'снизу свободно': (0, 1),
            'слева свободно': (-1, 0),
            'справа свободно': (1, 0)
        }

        if cond not in directions:
            raise ValueError(f"Неподдерживаемое условие: {condition}")

        dx, dy = directions[cond]

        result = self.is_valid_move(x + dx, y + dy)

        # Если было отрицание, инвертируем результат
        return not result if negate else result

    def handle_loop(self, condition, body):
        """Обработка циклов"""
        while self.check_condition(condition):
            for cmd in body:
                self.process_command(cmd)
                self.update()
                self.after(150)

    def handle_condition(self, condition, body):
        """Обработка условий"""
        if self.check_condition(condition):
            for cmd in body:
                self.process_command(cmd)
                self.update()
                self.after(150)

    def reset_maze(self):
        """Сброс состояния"""
        self.generate_maze()
        self.robot_pos = (2, self.grid_size//2 + 1)
        self.robot_path = set()
        self.code_editor.delete("1.0", END)  # Очистка текстового поля
        self.draw_maze()

    def on_mouse_scroll(self, event):
        """Прокрутка колесом мыши"""
        self.main_canvas.yview_scroll(int(-1*(event.delta/120)), "units")

    def update(self):
        """Обновление интерфейса"""
        self.draw_maze()
        self.update_idletasks()

    def check_solution(self):
        """Проверка условий с исключением специфических координат"""
        try:
            # Координаты стен
            h_y = self.grid_size // 2
            v_x = max(x for x, _ in self.walls['vertical'])

            # Целевые клетки (исключая проблемные координаты)
            target_cells = {
                (x, h_y + 1) for x in range(v_x) 
                if (x, h_y + 1) not in {(18, 23), (0, 13), (18, 22), (1, 13), (18, 24)}
            }.union({
                (v_x - 1, y) for y in range(h_y + 1, self.grid_size)
                if (v_x - 1, y) not in {(18, 23), (0, 13), (18, 22), (1, 13), (18, 24)}
            })

            # Удаление проходов
            target_cells.discard((self.passages['horizontal'], h_y + 1))
            target_cells.discard((v_x - 1, self.passages['vertical']))

            # Проверка выполнения
            missing = target_cells - self.robot_path
            
            if not missing:
                self.show_result("✓ Все условия выполнены!", "green")
            else:
                print(f"Осталось закрасить: {missing}")
                self.show_result("✗ Требуется доработка", "red")

        except Exception as e:
            print(f"Ошибка проверки: {str(e)}")
            self.show_result("✗ Критическая ошибка", "red")

    def show_result(self, message, color):
        """Отображение результата по центру экрана"""
        result_window = Toplevel(self)
        result_window.geometry("400x100")
        result_window.title("Результат")
        result_window.configure(bg='white')
        result_window.attributes("-topmost", True)
        
        # Центрирование окна
        screen_width = result_window.winfo_screenwidth()
        screen_height = result_window.winfo_screenheight()
        x = (screen_width // 2) - 200
        y = (screen_height // 2) - 50
        result_window.geometry(f"+{x}+{y}")
        
        Label(result_window, text=message, font=('Arial', 14), bg='white', fg=color).pack(pady=20)
        Button(result_window, text="OK", command=result_window.destroy, bg='#5DADE2', fg='white').pack()

# ================== ВСПОМОГАТЕЛЬНЫЕ ФУНКЦИИ ==================

def find_max_edge(matrix):
    G = nx.DiGraph()
    cities = ['А','Б','В','Г','Д','Е']
    
    for i in range(len(cities)):
        for j in range(len(cities)):
            if matrix[i][j] > 0:
                G.add_edge(cities[i], cities[j], weight=matrix[i][j])
    
    try:
        path = nx.shortest_path(G, 'А', 'Е', weight='weight')
        return max([G[u][v]['weight'] for u,v in zip(path[:-1], path[1:])])
    except:
        return np.max(matrix)

def calculate_b(start, end):
    current = start + 6  # +3 дважды
    required = end - 6   # +3 дважды после деления
    if required <= 0:
        return 1  # Предотвращение деления на 0
    return (start + 6) // required

# ================== ОКНО ТЕОРИИ ==================
class TheoryWindow(Toplevel):
    def __init__(self, parent, selected_tab=0):
        super().__init__(parent)
        self.title("Теоретические материалы")
        self.geometry("1400x900")
        self.configure(bg='white')
        
        self.notebook = ttk.Notebook(self)
        self.create_tabs()
        self.notebook.pack(expand=1, fill=BOTH)
        self.notebook.select(selected_tab)
        
        Button(self, text="Закрыть", command=self.destroy,
              font=('Arial', 12), bg='#EC7063', fg='white').pack(pady=10)

    def create_tabs(self):
        tabs = [
            ('1. Количественные параметры', self.task1_tab),
            ('2. Кодирование', self.task2_tab),
            ('3. Логика', self.task3_tab),
            ('4. Матрицы', self.task4_tab),
            ('5. Исполнители', self.task5_tab),
            ('6. Условия', self.task6_tab),
            ('7. URL', self.task7_tab),
            ('8. Поиск', self.task8_tab),
            ('9. Графы', self.task9_tab),
            ('10. Системы счисления', self.task10_tab),
            ('11. Поиск в тексте', self.task11_tab),
            ('12. Файловая система', self.task12_tab),
            ('13. Офисные пакеты', self.task13_tab),
            ('14. Электронные таблицы', self.task14_tab),
            ('15. Робот', self.task15_tab),
            ('16. Программирование', self.task16_tab)
        ]
        
        for name, method in tabs:
            frame = Frame(self.notebook, bg='white')
            method(frame)
            self.notebook.add(frame, text=name)

    def task1_tab(self, frame):
        content = """
Задание 1: Расчет информационного объема

Условие:
Статья содержит N страниц, на каждой странице M строк, 
в каждой строке K символов. Каждый символ кодируется 2 байтами. 

Формула:
Объем (КБ) = (N × M × K × 2) / 1024

Пример решения:
Дано:
- Страниц: 50
- Строк на странице: 40 
- Символов в строке: 60

Решение:
1. Общее количество символов: 50 × 40 × 60 = 120000
2. Объем в байтах: 120000 × 2 = 240000
3. Перевод в КБ: 240000 / 1024 ≈ 234.375

Ответ: 234 КБ"""
        self.create_text_content(frame, content)

    def task2_tab(self, frame):
        content = """
Задание 2: Кодирование сообщений

Условие:
Дана таблица кодов Морзе:
А: .- 
У: ..-
Ж: ...-
Х: ....

Пример сообщения: .-..-....-..

Алгоритм решения:
1. Разделить код на отдельные буквы
2. Сопоставить каждую последовательность с таблицей
3. Посчитать количество распознанных символов

Пример решения:
Код: .-..-....-..
Разбиение:
.- → А
..- → У
...- → Ж
.... → Х
..- → У

Итого: 5 символов"""
        self.create_text_content(frame, content)

    def task3_tab(self, frame):
        content = """
Задание 3: Логические выражения

Условие:
Найти наибольшее X, для которого истинно:
¬(X ≤ A) ∧ ¬(Y > B)

Алгоритм:
1. ¬(X ≤ A) → X > A
2. ¬(Y > B) → Y ≤ B
3. Максимальное X = A

Пример:
Дано: ¬(X ≤ 10) ∧ ¬(Y > 20)
Решение:
X > 10 и Y ≤ 20
Максимальное целое X = 10"""
        self.create_text_content(frame, content)

    def task4_tab(self, frame):
        content = """
Задание 4: Анализ матриц

Условие:
Дана матрица 6x6 с длинами дорог между пунктами.
Найти максимальный участок кратчайшего пути.

Алгоритм:
1. Построить граф по матрице
2. Найти все возможные пути
3. Выбрать максимальное значение в путях

Пример матрицы:
   А Б В Г Д Е
А 0 3 0 2 0 0
Б 0 0 5 0 4 0
В 0 0 0 0 0 7
Г 0 0 3 0 0 0
Д 0 0 0 0 0 2
Е 0 0 0 0 0 0

Кратчайший путь А-Б-Д-Е: 3+4+2=9
Максимальный участок: 4"""
        self.create_text_content(frame, content)

    def task5_tab(self, frame):
        content = """
Задание 5: Исполнитель Омега

Условие:
Исполнитель имеет команды:
1. Прибавить 3
2. Разделить на b

Алгоритм:
1. Восстановить последовательность операций
2. Составить уравнение

Пример:
Программа 11211 для числа 25 → 10
Решение:
25 +3 +3 /b +3 +3 = 10
Уравнение: (25+3+3)/b +3+3 = 10
(31)/b +6 = 10 → b = 31/4 = 7.75 → 7 (целое)"""
        self.create_text_content(frame, content)

    def task6_tab(self, frame):
        content = """
Задание 6: Условные операторы

Условие:
Определить количество выводов 'YES' для набора данных

Алгоритм:
1. Проверить каждую пару (s, t)
2. Подсчитать выполнение условия s < 9 or t < 9

Пример данных:
(9,9) → NO
(9,10) → NO
(8,5) → YES
(11,6) → YES
(-11,10) → YES

Ответ: 3"""
        self.create_text_content(frame, content)

    def task7_tab(self, frame):
        content = """
Задание 7: Построение URL

Правила:
- Начинается с протокола
- Домен/путь/файл

Пример:
Части: [http://, site.com, blog, post.html]
Результат: http://site.com/blog/post.html

Ошибки:
- Неверный порядок частей
- Отсутствие обязательных элементов"""
        self.create_text_content(frame, content)

    def task8_tab(self, frame):
        content = """
Задание 8: Поисковые запросы

Формула:
A = (A|B) + (A&B) - B

Пример:
Дано:
A|B = 3000
A&B = 500
B = 1500

Решение:
A = 3000 + 500 - 1500 = 2000

Объяснение:
Формула включений-исключений"""
        self.create_text_content(frame, content)

    def task9_tab(self, frame):
        content = """
Задание 9: Анализ графов

Условие:
Найти количество путей из А в И на фиксированном графе

Фиксированная схема:"""
        lbl = Label(frame, text=content, font=('Arial', 14), 
                   wraplength=1300, justify=LEFT, bg='white')
        lbl.pack(pady=10)
        
        # Фиксированный граф
        G = nx.DiGraph()
        edges = [('А','Б'), ('А','Г'), ('Б','В'), ('Б','Д'),
                ('В','Ж'), ('Г','В'), ('Г','Е'), ('Д','Ж'),
                ('Е','З'), ('Ж','И'), ('З','И')]
        G.add_edges_from(edges)
        
        fig = Figure(figsize=(8,6))
        ax = fig.add_subplot(111)
        pos = nx.spring_layout(G, seed=42)
        nx.draw(G, pos, ax=ax, with_labels=True, node_size=800,
               node_color='#E8F8FF', arrowsize=20)
        
        canvas = FigureCanvasTkAgg(fig, frame)
        canvas.draw()
        canvas.get_tk_widget().pack()
        
        solution = """
Пути:
1. А → Б → В → Ж → И
2. А → Б → Д → Ж → И
3. А → Г → В → Ж → И
4. А → Г → Е → З → И

Всего: 4 пути"""
        Label(frame, text=solution, font=('Arial', 14),
             bg='white').pack(pady=10)

    def task10_tab(self, frame):
        content = """
Задание 10: Системы счисления

Алгоритм:
1. Перевести все числа в десятичную систему
2. Сравнить результаты

Пример:
41₈ = 4×8¹ + 1×8⁰ = 33
77₈ = 7×8¹ + 7×8⁰ = 63
10001₂ = 1×2⁴ + 0×2³ + 0×2² + 0×2¹ + 1×2⁰ = 17

Минимальное: 17"""
        self.create_text_content(frame, content)

    def task11_tab(self, frame):
        content = """
Задание 11: Поиск информации в текстах

Алгоритм:
1. Открыть текстовый файл из архива
2. Использовать поиск (Ctrl+F) по ключевым словам
3. Прочитать контекст вокруг найденного упоминания

Пример для "Евгений Базаров":
1. В архиве DEMO-12 находим роман "Отцы и дети"
2. Ищем "Базаров" -> находим: "Евгений Базаров, нигилист..."
3. В соседних предложениях находим имя

Типичные ошибки:
- Поиск по неверному формату файла
- Не учитываются сокращения (Евг. вместо Евгений)"""
        self.create_text_content(frame, content)

    def task12_tab(self, frame):
        content = """
Задание 12: Работа с файловой системой

Алгоритм:
1. Распаковать архив в отдельную папку
2. Использовать поиск по расширению:
   - В Windows: *.doc в поле поиска проводника
   - В Linux: find . -name "*.doc"

Пример структуры:
Task12/
├── Sub1/
│   ├── file1.doc
│   └── file2.txt
└── Sub2/
    └── file3.doc

Ответ: 2 файла .doc

Важно:
- Учитывать вложенные папки
- Проверять регистр расширений (.DOC и .doc)"""
        self.create_text_content(frame, content)

    def task13_tab(self, frame):
        content = """
Задание 13: Создание презентации/документа

Требования к презентации:
1. Единый стиль (шрифты, цвета, фон)
2. Заголовки на каждом слайде
3. Использование изображений из материалов

Пример структуры слайдов:
1. Титульный: Название и автор
2. Основной: Таблица/диаграмма
3. Заключительный: Выводы

Требования к документу:
- Отступ первой строки: 1 см
- Выравнивание по ширине
- Границы таблицы толщиной 1 пт"""
        self.create_text_content(frame, content)

    def task14_tab(self, frame):
        content = """
Задание 14: Анализ в электронных таблицах

Функции:
1. СЧЁТЕСЛИМН() - для подсчета по условиям
   Пример: =СЧЁТЕСЛИМН(A:A,"СВ";C:C,"математика")

2. СРЗНАЧЕСЛИ() - для среднего по условию
   Пример: =СРЗНАЧЕСЛИ(A:A,"Ю";D:D)

3. Диаграмма:
- Выделить данные
- Вставка → Круговая диаграмма
- Настройка подписей данных"""
        self.create_text_content(frame, content)

    def task15_tab(self, frame):
        content = """
Задание 15: Программирование Робота

Фиксированное поле:
┌───────────────┐
│ ■ ■ ■ □ □ □ □ │
│ □ □ ■ □ □ □ □ │
│ □ □ ■ ■ ■ □ □ │ ← Горизонтальная стена
│ □ □ □ □ │ □ □ │
│ □ □ □ □ │ □ □ │ ← Вертикальная стена
└───────────────┘

Алгоритм:
1. Двигаться вправо до стены
2. Закрашивать клетки слева от вертикальной стены
3. Проверять свободные клетки сверху

Пример программы:
вправо
нц пока сверху свободно
закрасить
вверх
кц
вниз
закрасить"""
        self.create_text_content(frame, content)

    def task16_tab(self, frame):
        content = """
Задание 16: Анализ температур

Пример входных данных:
6
5 -2 10 0 7 -3

Алгоритм решения:
1. Фильтрация положительных температур
2. Расчет среднего значения
3. Подсчет количества дней

Решение на Python:
n = int(input())
temps = [int(input()) for _ in range(n)]
positive = [t for t in temps if t > 0]
avg = sum(positive)/len(positive)
print(avg, len(positive))"""
        self.create_text_content(frame, content)

    def create_text_content(self, frame, text):
        lbl = Label(frame, text=text, font=('Arial', 14), 
                   wraplength=1300, justify=LEFT, bg='white')
        lbl.pack(pady=20, padx=20, anchor='w')
# ================== ОКНО ЗАДАНИЯ ==================
class TaskWindow(Toplevel):
    def __init__(self, parent, task_data):
        super().__init__(parent)
        self.title(f"Задание {task_data['type']}")
        self.geometry("1000x700")
        self.configure(bg='white')
        self.task_data = task_data
        self.result_label = None
        self.fade_alpha = 100
        self.fade_step = 5

        # Основной контейнер с прокруткой
        main_frame = Frame(self, bg='white')
        main_frame.pack(fill=BOTH, expand=True)

        # Настройка прокрутки
        canvas = Canvas(main_frame, bg='white', highlightthickness=0)
        scrollbar = ttk.Scrollbar(main_frame, orient="vertical", command=canvas.yview)
        self.scrollable_frame = Frame(canvas, bg='white')
        canvas.bind_all("<MouseWheel>", lambda e: canvas.yview_scroll(int(-1*(e.delta/120)), "units"))

        canvas.configure(yscrollcommand=scrollbar.set)
        canvas.pack(side=LEFT, fill=BOTH, expand=True)
        scrollbar.pack(side=RIGHT, fill=Y)
        canvas.create_window((0, 0), window=self.scrollable_frame, anchor="nw")

        # Обработчик изменения размера
        self.scrollable_frame.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))

        # Заголовок задания
        Label(self.scrollable_frame,
             text=f"Задание {task_data['type']}",
             font=('Arial', 16, 'bold'),
             bg='white').pack(pady=15)

        # Условие задачи
        condition_frame = Frame(self.scrollable_frame, bg='white')
        condition_frame.pack(fill=X, padx=20, pady=10)

        for line in task_data['condition'].split('\n'):
            Label(condition_frame,
                 text=line.strip(),
                 font=('Arial', 14),
                 wraplength=900,
                 justify=LEFT,
                 bg='white').pack(anchor='w', pady=2)

        # Специфические элементы
        if task_data['type'] == 'matrix':
            self.show_matrix(task_data['matrix'])
        elif task_data['type'] == 'graph':
            self.show_graph(task_data['graph'])
        elif task_data['type'] == 'morse':
            self.show_morse_table(task_data['table'])
            Label(self.scrollable_frame,
                 text=f"Код: {task_data['encoded']}",
                 font=('Courier', 14),
                 bg='white').pack(pady=10)
        elif task_data['type'] == 'spreadsheet':
            self.show_spreadsheet(task_data)
        
        # Добавление изображения
        if 'image' in task_data:
            try:
                img = Image.open(task_data['image'])
                img = img.resize((400, 300), Image.Resampling.LANCZOS)
                photo = ImageTk.PhotoImage(img)
                img_label = Label(self.scrollable_frame, image=photo)
                img_label.image = photo
                img_label.pack(pady=10)
            except Exception as e:
                print(f"Ошибка загрузки изображения: {e}")

        # Блок ответа
        answer_frame = Frame(self.scrollable_frame, bg='white')
        answer_frame.pack(pady=20)

        self.entry = Entry(answer_frame,
                          font=('Arial', 14),
                          width=25,
                          justify=CENTER,
                          relief='solid',
                          borderwidth=2)
        self.entry.pack(side=LEFT, padx=10)

        Button(answer_frame,
              text="Проверить",
              command=lambda: self.check_answer(task_data['answer']),
              font=('Arial', 12),
              bg='#58D68D',
              fg='white',
              padx=15,
              pady=5).pack(side=LEFT)

        # Кнопка закрытия
        Button(self.scrollable_frame,
              text="Закрыть",
              command=self.destroy,
              font=('Arial', 12),
              bg='#EC7063',
              fg='white',
              padx=20,
              pady=5).pack(pady=15)

    def show_matrix(self, matrix):
        matrix_frame = Frame(self.scrollable_frame, bg='white')
        matrix_frame.pack(pady=15)

        cities = ['А','Б','В','Г','Д']
        table_frame = Frame(matrix_frame, bg='white')
        table_frame.pack()

        # Заголовки
        for j, city in enumerate(cities):
            Label(table_frame,
                text=city,
                width=7,
                relief='ridge',
                font=('Arial', 12, 'bold'),
                bg='#E0E0E0').grid(row=0, column=j+1, sticky='nsew')

        # Тело таблицы
        for i in range(matrix.shape[0]):
            Label(table_frame,
                text=cities[i],
                width=7,
                relief='ridge',
                font=('Arial', 12),
                bg='#E0E0E0').grid(row=i+1, column=0, sticky='nsew')

            for j in range(matrix.shape[1]):
                bg = '#FFFFFF' if i != j else '#F0F0F0'
                # Если значение в матрице равно 0, отображаем пустую строку
                cell_value = str(matrix[i][j]) if matrix[i][j] != 0 else ""
                Label(table_frame,
                    text=cell_value,
                    width=7,
                    relief='ridge',
                    font=('Arial', 12),
                    bg=bg).grid(row=i+1, column=j+1, sticky='nsew')


    def show_graph(self, figure):
        graph_frame = Frame(self.scrollable_frame, bg='white')
        graph_frame.pack(pady=15, fill=BOTH, expand=True)

        canvas = FigureCanvasTkAgg(figure, graph_frame)
        canvas.draw()
        canvas.get_tk_widget().pack(fill=BOTH, expand=True)

    def show_morse_table(self, table):
        table_frame = Frame(self.scrollable_frame, bg='white')
        table_frame.pack(pady=15)

        for i, (char, code) in enumerate(table.items()):
            Label(table_frame,
                 text=f"{char}: {code}",
                 font=('Courier', 14),
                 relief='ridge',
                 padx=15,
                 pady=5,
                 bg='white').grid(row=i//2, column=i%2, padx=5, pady=5)

    def show_spreadsheet(self, task_data):
        file_frame = Frame(self.scrollable_frame, bg='white')
        file_frame.pack(pady=10)
        
        # Кнопка скачивания файла
        Button(file_frame, 
            text="Скачать файл", 
            command=lambda: os.startfile(task_data['file']),
            font=('Arial', 12),
            bg='#5DADE2').pack(side=LEFT)
        
        Label(file_frame, 
            text=task_data['file'],
            font=('Arial', 12),
            bg='white').pack(side=LEFT, padx=10)
        
        # Отображение первых 5 строк таблицы
        table_frame = Frame(self.scrollable_frame, bg='white')
        table_frame.pack(pady=15)
        
        # Заголовки
        headers = ['A', 'B', 'C', 'D']
        for j, header in enumerate(headers):
            Label(table_frame, 
                text=header,
                width=15,
                relief='ridge',
                font=('Arial', 12, 'bold'),
                bg='#E0E0E0').grid(row=0, column=j, sticky='nsew')
        
        # Данные
        df = pd.read_excel(task_data['file'])
        for i in range(5):
            for j, col in enumerate(headers):
                Label(table_frame, 
                    text=df.iloc[i, j],
                    width=15,
                    relief='ridge',
                    font=('Arial', 12),
                    bg='white').grid(row=i+1, column=j, sticky='nsew')

    def check_answer(self, correct):
        user_answer = self.entry.get().strip()
        is_correct = user_answer == str(correct)

        if is_correct:
            self.show_result("✓ Правильно!", "green")
        else:
            self.show_result(f"✗ Ошибка! Правильно: {correct}", "red")

    def show_result(self, message, color):
        """Отображение результата по центру экрана"""
        result_window = Toplevel(self)
        result_window.geometry("400x100")
        result_window.title("Результат")
        result_window.configure(bg='white')
        result_window.attributes("-topmost", True)
        
        # Центрирование окна
        screen_width = result_window.winfo_screenwidth()
        screen_height = result_window.winfo_screenheight()
        x = (screen_width // 2) - 200
        y = (screen_height // 2) - 50
        result_window.geometry(f"+{x}+{y}")
        
        Label(result_window, text=message, font=('Arial', 14), bg='white', fg=color).pack(pady=20)
        Button(result_window, text="OK", command=result_window.destroy, bg='#5DADE2', fg='white').pack()
    
    
# ================== ГЛАВНОЕ ОКНО ==================
class OGEApp(Tk):
    def __init__(self):
        super().__init__()
        self.title("ОГЭ 2025 - Тренажер")
        self.geometry("1000x600")
        self.configure(bg='white')
        self.tasks = self.generate_all_tasks()
        self.current_windows = []
        
        # Главный контейнер
        self.main_container = Frame(self, bg='white')
        self.main_container.pack(fill=BOTH, expand=True)
        
        self.add_regenerate_button()
        self.show_main_menu()

    def add_regenerate_button(self):
        """Добавляет кнопку перегенерации заданий"""
        self.regenerate_btn = Button(self, 
                                    text="♻ Новые задания", 
                                    command=self.regenerate_tasks,
                                    font=('Arial', 12),
                                    bg='#5DADE2',
                                    fg='white',
                                    padx=10,
                                    pady=5)
        self.regenerate_btn.pack(side=BOTTOM, pady=(0, 5))

    def regenerate_tasks(self):
        """Перегенерировать все задания"""
        if messagebox.askyesno("Подтверждение", 
                             "Все открытые задания будут закрыты. Продолжить?"):
            # Закрыть все открытые окна заданий
            for window in self.current_windows.copy():
                if window.winfo_exists():
                    window.destroy()
            self.current_windows.clear()
            
            # Перегенерировать задания
            self.tasks = self.generate_all_tasks()
            
            # Обновить главное меню
            self.show_main_menu()
            messagebox.showinfo("Обновлено", "Все задания успешно перегенерированы!")

    def show_main_menu(self):
        """Показать главное меню"""
        self.clear_container()
        main_frame = Frame(self.main_container, bg='white')
        main_frame.place(relx=0.5, rely=0.5, anchor=CENTER)
        
        Label(main_frame, 
             text="ОГЭ 2025 - Тренажер", 
             font=('Arial', 20, 'bold'), 
             bg='white').pack(pady=20)
        
        btn_frame = Frame(main_frame, bg='white')
        Button(btn_frame, 
              text="Практика", 
              command=self.show_practice,
              font=('Arial', 14), 
              width=15, 
              height=2).pack(side=LEFT, padx=10)
        Button(btn_frame, 
              text="Теория", 
              command=self.show_theory,
              font=('Arial', 14), 
              width=15, 
              height=2).pack(side=LEFT, padx=10)
        btn_frame.pack(pady=20)

    def show_practice(self):
        """Показать сетку заданий"""
        self.clear_container()
        container = Frame(self.main_container, bg='white')
        container.pack(fill=BOTH, expand=True, padx=20, pady=20)
        
        # Сетка заданий 4x4
        grid_frame = Frame(container, bg='white')
        grid_frame.pack()
        
        for i in range(16):
            row = i // 4
            col = i % 4
            btn_text = f"Задание {i+1}"
            
            # Особые обработчики для заданий 15 и 16
            if i == 14:
                cmd = self.open_task15
            elif i == 15:
                cmd = self.open_task16
            else:
                cmd = lambda idx=i: self.open_task(idx)
            
            btn = Button(grid_frame, 
                        text=btn_text, 
                        command=cmd,
                        font=('Arial', 12), 
                        width=15, 
                        height=2)
            btn.grid(row=row, column=col, padx=10, pady=10)
        
        # Кнопка назад
        Button(container, 
              text="Назад", 
              command=self.show_main_menu,
              font=('Arial', 12), 
              bg='#EC7063').pack(side=BOTTOM, pady=10)

    def show_theory(self):
        """Открыть окно теории"""
        TheoryWindow(self)

    def open_task(self, task_index):
        """Открыть конкретное задание"""
        if 0 <= task_index < len(self.tasks):
            task_window = TaskWindow(self, self.tasks[task_index])
            self.current_windows.append(task_window)
            task_window.protocol("WM_DELETE_WINDOW", 
                               lambda: self.on_task_window_close(task_window))

    def on_task_window_close(self, window):
        """Обработчик закрытия окна задания"""
        if window in self.current_windows:
            self.current_windows.remove(window)
        window.destroy()

    def open_task15(self):
        """Специальный обработчик для задания 15"""
        task_window = RobotTaskWindow(self)
        self.current_windows.append(task_window)
        task_window.protocol("WM_DELETE_WINDOW", 
                           lambda: self.on_task_window_close(task_window))

    def open_task16(self):
        """Специальный обработчик для задания 16"""
        task_data = generate_task16()
        task_window = Task16Window(self, task_data)
        self.current_windows.append(task_window)
        task_window.protocol("WM_DELETE_WINDOW", 
                           lambda: self.on_task_window_close(task_window))

    def generate_all_tasks(self):
        """Генерация всех 16 заданий"""
        return [
            generate_task1(),
            generate_task2(),
            generate_task3(),
            generate_task4(),
            generate_task5(),
            generate_task6(),
            generate_task7(),
            generate_task8(),
            generate_task9(),
            generate_task10(),
            generate_task11(),
            generate_task12(),
            generate_task13(),
            generate_task14(),
            generate_task15(),
            generate_task16()
        ]

    def clear_container(self):
        """Очистка главного контейнера"""
        for widget in self.main_container.winfo_children():
            widget.destroy()

    def mainloop(self):
        """Переопределенный mainloop с очисткой"""
        super().mainloop()
        # Очистка временных файлов при закрытии
        for task in self.tasks:
            if 'file' in task and os.path.exists(task['file']):
                try:
                    os.remove(task['file'])
                except Exception as e:
                    print(f"Ошибка удаления файла: {e}")

if __name__ == "__main__":
    app = OGEApp()
    app.mainloop()